"""
Attachment Parser - Extract text from PDF, Word, Excel files
Supports OCR and Vision AI (Claude, Gemini) for image analysis
"""

import os
import base64
from pathlib import Path
from typing import Optional, Dict, Any, Literal

from .config import get_config


# Check available parsers
PYMUPDF_AVAILABLE = False
DOCX_AVAILABLE = False
OPENPYXL_AVAILABLE = False
TESSERACT_AVAILABLE = False
ANTHROPIC_AVAILABLE = False
GOOGLE_GENAI_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    pass

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    pass

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

try:
    import pytesseract
    from PIL import Image
    # Set Tesseract path on Windows
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    TESSERACT_AVAILABLE = True
except ImportError:
    pass

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    pass

try:
    import google.generativeai as genai
    GOOGLE_GENAI_AVAILABLE = True
except ImportError:
    pass


# Vision AI prompt for image analysis
VISION_PROMPT = """Analyze this image and describe its contents in detail.
If it's a technical drawing, floor plan, or architectural document:
- Describe what type of drawing it is
- List key elements, labels, dimensions visible
- Note any text or annotations

If it's a photo or screenshot:
- Describe what is shown
- Note any text visible

Provide a concise but comprehensive description that would help someone search for this image later.
Respond in the same language as any text visible in the image (Korean or English)."""


def _get_image_base64(file_path: str) -> tuple[str, str]:
    """Read image and return base64 encoded data with media type"""
    ext = Path(file_path).suffix.lower()
    media_type_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    media_type = media_type_map.get(ext, "image/png")

    with open(file_path, "rb") as f:
        image_data = base64.standard_b64encode(f.read()).decode("utf-8")

    return image_data, media_type


def analyze_image_with_claude(file_path: str, api_key: Optional[str] = None) -> Dict[str, Any]:
    """Analyze image using Claude Vision API"""
    if not ANTHROPIC_AVAILABLE:
        return {"success": False, "error": "anthropic package not installed", "text": ""}

    if api_key is None:
        api_key = os.environ.get("ANTHROPIC_API_KEY")

    if not api_key:
        return {"success": False, "error": "ANTHROPIC_API_KEY not set", "text": ""}

    try:
        image_data, media_type = _get_image_base64(file_path)

        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_data,
                            },
                        },
                        {
                            "type": "text",
                            "text": VISION_PROMPT,
                        }
                    ],
                }
            ],
        )

        text = message.content[0].text
        return {
            "success": True,
            "text": text,
            "vision": "claude",
            "model": "claude-sonnet-4-20250514",
            "tokens_used": message.usage.input_tokens + message.usage.output_tokens,
        }
    except Exception as e:
        return {"success": False, "error": str(e), "text": ""}


def analyze_image_with_gemini(file_path: str, api_key: Optional[str] = None) -> Dict[str, Any]:
    """Analyze image using Gemini Vision API"""
    if not GOOGLE_GENAI_AVAILABLE:
        return {"success": False, "error": "google-generativeai package not installed", "text": ""}

    if api_key is None:
        api_key = os.environ.get("GOOGLE_API_KEY")

    if not api_key:
        return {"success": False, "error": "GOOGLE_API_KEY not set", "text": ""}

    try:
        genai.configure(api_key=api_key)

        # Load image
        from PIL import Image
        img = Image.open(file_path)

        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content([VISION_PROMPT, img])

        text = response.text
        return {
            "success": True,
            "text": text,
            "vision": "gemini",
            "model": "gemini-2.0-flash",
        }
    except Exception as e:
        return {"success": False, "error": str(e), "text": ""}


class AttachmentParser:
    """Parse various attachment types and extract text"""

    def __init__(self):
        self.config = get_config()
        self.config.ensure_directories()

    def get_supported_types(self) -> Dict[str, bool]:
        """Get supported file types and their availability"""
        return {
            ".pdf": PYMUPDF_AVAILABLE,
            ".docx": DOCX_AVAILABLE,
            ".xlsx": OPENPYXL_AVAILABLE,
            ".txt": True,
            ".png": TESSERACT_AVAILABLE,
            ".jpg": TESSERACT_AVAILABLE,
            ".jpeg": TESSERACT_AVAILABLE,
        }

    def get_vision_availability(self) -> Dict[str, bool]:
        """Get Vision AI availability"""
        return {
            "claude": ANTHROPIC_AVAILABLE,
            "gemini": GOOGLE_GENAI_AVAILABLE,
        }

    def parse_file(self, file_path: str, vision_provider: Optional[Literal["claude", "gemini", "ocr"]] = None) -> Dict[str, Any]:
        """
        Parse a file and extract text

        Args:
            file_path: Path to the file
            vision_provider: For images, use "claude", "gemini", or "ocr" (default: "ocr")

        Returns:
            Dict with 'success', 'text', and optionally 'error'
        """
        if not os.path.exists(file_path):
            return {"success": False, "error": "File not found", "text": ""}

        ext = Path(file_path).suffix.lower()

        try:
            if ext == ".pdf":
                return self._parse_pdf(file_path)
            elif ext == ".docx":
                return self._parse_docx(file_path)
            elif ext == ".xlsx":
                return self._parse_xlsx(file_path)
            elif ext == ".txt":
                return self._parse_txt(file_path)
            elif ext in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
                return self._parse_image(file_path, vision_provider)
            else:
                return {"success": False, "error": f"Unsupported file type: {ext}", "text": ""}
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}

    def _parse_pdf(self, file_path: str) -> Dict[str, Any]:
        """Extract text from PDF using PyMuPDF"""
        if not PYMUPDF_AVAILABLE:
            return {"success": False, "error": "PyMuPDF not installed", "text": ""}

        try:
            doc = fitz.open(file_path)
            text_parts = []

            for page_num, page in enumerate(doc):
                page_text = page.get_text()
                if page_text.strip():
                    text_parts.append(f"[Page {page_num + 1}]\n{page_text}")

            doc.close()

            if not text_parts:
                return {"success": False, "error": "No text found in PDF", "text": ""}

            return {
                "success": True,
                "text": "\n\n".join(text_parts),
                "pages": len(text_parts),
            }
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}

    def _parse_docx(self, file_path: str) -> Dict[str, Any]:
        """Extract text from Word document"""
        if not DOCX_AVAILABLE:
            return {"success": False, "error": "python-docx not installed", "text": ""}

        try:
            doc = docx.Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

            if not paragraphs:
                return {"success": False, "error": "No text found in document", "text": ""}

            return {
                "success": True,
                "text": "\n\n".join(paragraphs),
                "paragraphs": len(paragraphs),
            }
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}

    def _parse_xlsx(self, file_path: str) -> Dict[str, Any]:
        """Extract text from Excel spreadsheet"""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "openpyxl not installed", "text": ""}

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = [f"[Sheet: {sheet_name}]"]

                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                    if row_values:
                        sheet_text.append(" | ".join(row_values))

                if len(sheet_text) > 1:  # More than just header
                    text_parts.append("\n".join(sheet_text))

            wb.close()

            if not text_parts:
                return {"success": False, "error": "No data found in spreadsheet", "text": ""}

            return {
                "success": True,
                "text": "\n\n".join(text_parts),
                "sheets": len(text_parts),
            }
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}

    def _parse_txt(self, file_path: str) -> Dict[str, Any]:
        """Read text file"""
        try:
            # Try different encodings
            for encoding in ["utf-8", "cp1252", "latin-1"]:
                try:
                    with open(file_path, "r", encoding=encoding) as f:
                        text = f.read()
                    return {"success": True, "text": text, "encoding": encoding}
                except UnicodeDecodeError:
                    continue

            return {"success": False, "error": "Could not decode text file", "text": ""}
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}

    def _parse_image(self, file_path: str, vision_provider: Optional[str] = None) -> Dict[str, Any]:
        """Extract text/description from image using OCR or Vision AI

        Args:
            file_path: Path to image file
            vision_provider: "claude", "gemini", or "ocr" (default)
        """
        # Use Vision AI if specified
        if vision_provider == "claude":
            return analyze_image_with_claude(file_path)
        elif vision_provider == "gemini":
            return analyze_image_with_gemini(file_path)

        # Default to OCR
        if not TESSERACT_AVAILABLE:
            return {"success": False, "error": "pytesseract not installed", "text": ""}

        try:
            img = Image.open(file_path)

            # Use Korean + English languages
            text = pytesseract.image_to_string(img, lang="kor+eng")

            if not text.strip():
                return {"success": False, "error": "No text found in image", "text": ""}

            return {"success": True, "text": text, "ocr": True}
        except Exception as e:
            return {"success": False, "error": str(e), "text": ""}


# Singleton
_parser: Optional[AttachmentParser] = None


def get_attachment_parser() -> AttachmentParser:
    """Get or create singleton AttachmentParser"""
    global _parser
    if _parser is None:
        _parser = AttachmentParser()
    return _parser


def extract_attachment_text(file_path: str) -> Optional[str]:
    """
    Convenience function to extract text from a file

    Returns:
        Extracted text, or None if extraction failed
    """
    parser = get_attachment_parser()
    result = parser.parse_file(file_path)

    if result.get("success"):
        return result.get("text")
    return None
